from flask import Flask, request, jsonify, send_file, render_template, make_response
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
import io, os, urllib.parse, tempfile, shutil, errno, logging, subprocess
from uuid import uuid4
from concurrent.futures import ThreadPoolExecutor
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional
import re

# Adobe PDF Services SDK imports - v4.2.0 compatible
ADOBE_AVAILABLE = False
try:
    from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
    from adobe.pdfservices.operation.pdf_services import PDFServices
    from adobe.pdfservices.operation.io.stream_asset import StreamAsset
    from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
    from adobe.pdfservices.operation.pdfjobs.jobs.export_pdf_job import ExportPDFJob
    from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_params import ExportPDFParams
    from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_target_format import ExportPDFTargetFormat
    from adobe.pdfservices.operation.pdfjobs.result.export_pdf_result import ExportPDFResult
    from adobe.pdfservices.operation.exception.exceptions import ServiceApiException, ServiceUsageException, SdkException
    ADOBE_AVAILABLE = True
    logging.info("Adobe PDF Services SDK v4.2.0 loaded successfully")
except ImportError as e:
    logging.warning(f"Adobe PDF Services SDK not available: {e}")
    logging.info("Service will use fallback image-based conversion")

app = Flask(__name__, template_folder="templates", static_folder="static")
logging.basicConfig(level=logging.INFO)
CORS(app, resources={r"/*": {"origins": [
    "http://localhost:5173",
    "https://77-tools.xyz",
    "https://www.77-tools.xyz",
    "https://popular-77.vercel.app"
], "expose_headers": ["Content-Disposition"], "methods": ["GET","POST","OPTIONS"], "allow_headers": ["Content-Type"]}})

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)

executor = ThreadPoolExecutor(max_workers=2)
JOBS = {}
current_job_id: Optional[str] = None

def safe_base_name(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename or "output"))[0]
    return base.replace("/", "").replace("\\", "").strip() or "output"

def set_progress(job_id, p, msg=None):
    info = JOBS.setdefault(job_id, {})
    info["progress"] = int(p)
    if msg is not None:
        info["message"] = msg

def send_download_memory(path: str, download_name: str, ctype: str):
    if not path or not os.path.exists(path):
        return jsonify({"error":"output file missing"}), 500
    with open(path, "rb") as f:
        data = f.read()
    resp = send_file(io.BytesIO(data), mimetype=ctype, as_attachment=True, download_name=download_name, conditional=False)
    resp.direct_passthrough = False
    resp.headers["Content-Length"] = str(len(data))
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{urllib.parse.quote(download_name)}"
    return resp

def _env(key: str, alt: list[str] = []):
    import os
    for k in [key, *alt]:
        v = os.environ.get(k)
        if v: return v
    raise KeyError(f"Missing env: {key}")

def adobe_context():
    """Adobe PDF Services SDK v4.2.0 자격 증명 설정"""
    from dotenv import load_dotenv
    load_dotenv()
    
    # ADOBE_* 또는 PDF_SERVICES_* 둘 다 허용
    client_id = os.environ.get("ADOBE_CLIENT_ID") or os.environ.get("PDF_SERVICES_CLIENT_ID")
    client_secret = os.environ.get("ADOBE_CLIENT_SECRET") or os.environ.get("PDF_SERVICES_CLIENT_SECRET")
    
    if not client_id or not client_secret:
        raise KeyError("Adobe credentials missing: ADOBE_CLIENT_ID / ADOBE_CLIENT_SECRET")
    
    # SDK v4.2.0에서는 ServicePrincipalCredentials만 사용
    creds = ServicePrincipalCredentials(
        client_id=client_id,
        client_secret=client_secret
    )
    return creds

def _export_via_adobe(in_pdf_path: str, target: str, out_path: str):
    """Adobe PDF Services SDK v4.2.0을 사용한 PDF 변환"""
    try:
        # 자격 증명 생성
        credentials = adobe_context()
        
        # PDF Services 클라이언트 생성
        pdf_services = PDFServices(credentials=credentials)
        
        # PDF 파일을 StreamAsset으로 업로드
        with open(in_pdf_path, 'rb') as file:
            input_stream = file.read()
        
        input_asset = pdf_services.upload(
            input_stream=input_stream, 
            mime_type=PDFServicesMediaType.PDF
        )
        
        # Export 파라미터 설정
        if target == "XLSX":
            target_format = ExportPDFTargetFormat.XLSX
        else:
            raise ValueError(f"Unsupported target format: {target}")
            
        export_pdf_params = ExportPDFParams(target_format=target_format)
        
        # ExportPDFJob 생성 및 실행
        export_pdf_job = ExportPDFJob(
            input_asset=input_asset,
            export_pdf_params=export_pdf_params
        )
        
        # 작업 제출 및 결과 대기
        location = pdf_services.submit(export_pdf_job)
        pdf_services_response = pdf_services.get_job_result(location, ExportPDFResult)
        
        # 결과 파일 다운로드 및 저장
        result_asset = pdf_services_response.get_result().get_asset()
        stream_asset = pdf_services.get_content(result_asset)
        
        # 기존 파일이 있으면 삭제
        if os.path.exists(out_path):
            os.remove(out_path)
            
        # 결과를 파일로 저장
        with open(out_path, 'wb') as output_file:
            output_file.write(stream_asset.get_input_stream())
            
        logging.info(f"Adobe SDK v4.2.0 변환 완료: {out_path}")
        
    except (ServiceApiException, ServiceUsageException, SdkException) as e:
        logging.error(f"Adobe SDK 변환 실패: {e}")
        raise
    except Exception as e:
        logging.error(f"Adobe SDK 변환 중 예상치 못한 오류: {e}")
        raise

def perform_xlsx_conversion_adobe(in_path: str, base_name: str):
    """Adobe SDK를 사용한 XLSX 변환"""
    final_name = f"{base_name}.xlsx"
    final_path = os.path.join(OUTPUTS_DIR, final_name)
    _export_via_adobe(in_path, "XLSX", final_path)
    return final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def perform_xlsx_conversion_adobe_with_sheets(in_path: str, base_name: str):
    """Adobe SDK를 사용한 XLSX 변환 후 각 페이지별 시트로 재구성"""
    import openpyxl
    
    # 먼저 Adobe SDK로 변환
    temp_path, temp_name, ctype = perform_xlsx_conversion_adobe(in_path, base_name)
    
    # PDF 페이지 수 확인
    doc = fitz.open(in_path)
    page_count = doc.page_count
    doc.close()
    
    # Adobe SDK 결과 파일 읽기
    adobe_wb = openpyxl.load_workbook(temp_path)
    adobe_ws = adobe_wb.active
    
    # 새로운 워크북 생성 (각 페이지별 시트)
    new_wb = Workbook()
    first_sheet = True
    
    # 전체 데이터를 페이지 수로 나누어 각 시트에 배치
    total_rows = adobe_ws.max_row
    rows_per_page = max(1, total_rows // page_count) if page_count > 0 else total_rows
    
    for page_num in range(1, page_count + 1):
        if first_sheet:
            ws = new_wb.active
            ws.title = f"Page_{page_num:02d}"
            first_sheet = False
        else:
            ws = new_wb.create_sheet(title=f"Page_{page_num:02d}")
        
        # 해당 페이지에 해당하는 행 범위 계산
        start_row = (page_num - 1) * rows_per_page + 1
        end_row = min(page_num * rows_per_page, total_rows) if page_num < page_count else total_rows
        
        # 데이터 복사
        row_offset = 0
        for row_num in range(start_row, end_row + 1):
            row_offset += 1
            for col_num in range(1, adobe_ws.max_column + 1):
                source_cell = adobe_ws.cell(row=row_num, column=col_num)
                target_cell = ws.cell(row=row_offset, column=col_num)
                
                # 값과 스타일 복사
                target_cell.value = source_cell.value
                if source_cell.font:
                    target_cell.font = source_cell.font
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment
    
    # 새로운 파일로 저장
    final_name = f"{base_name}.xlsx"
    final_path = os.path.join(OUTPUTS_DIR, final_name)
    if os.path.exists(final_path):
        os.remove(final_path)
    new_wb.save(final_path)
    
    # 임시 파일 정리
    try:
        os.remove(temp_path)
    except:
        pass
    
    return final_path, final_name, ctype

def has_extractable_text(page):
    """PDF 페이지에서 추출 가능한 텍스트가 있는지 확인"""
    text = page.get_text().strip()
    # 텍스트가 있고, 단순히 공백이나 특수문자만이 아닌 경우
    return len(text) > 10 and bool(re.search(r'[a-zA-Z가-힣0-9]', text))

def extract_text_to_excel(page, ws):
    """PDF 페이지에서 텍스트를 추출하여 Excel 셀에 배치"""
    # 텍스트 블록 추출 (위치 정보 포함)
    blocks = page.get_text("dict")["blocks"]
    
    # 텍스트 블록들을 y 좌표로 정렬 (위에서 아래로)
    text_blocks = []
    for block in blocks:
        if "lines" in block:  # 텍스트 블록인 경우
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        text_blocks.append({
                            "text": text,
                            "x": span["bbox"][0],
                            "y": span["bbox"][1],
                            "font_size": span["size"]
                        })
    
    # y 좌표로 정렬 후 x 좌표로 정렬
    text_blocks.sort(key=lambda x: (x["y"], x["x"]))
    
    # 행별로 그룹화 (비슷한 y 좌표끼리)
    rows = []
    current_row = []
    current_y = None
    tolerance = 5  # y 좌표 허용 오차
    
    for block in text_blocks:
        if current_y is None or abs(block["y"] - current_y) <= tolerance:
            current_row.append(block)
            current_y = block["y"] if current_y is None else current_y
        else:
            if current_row:
                rows.append(current_row)
            current_row = [block]
            current_y = block["y"]
    
    if current_row:
        rows.append(current_row)
    
    # Excel에 텍스트 배치
    for row_idx, row_blocks in enumerate(rows, 1):
        # 행 내에서 x 좌표로 정렬
        row_blocks.sort(key=lambda x: x["x"])
        
        # 열 위치 계산 (x 좌표 기반)
        col_positions = {}
        for i, block in enumerate(row_blocks):
            # x 좌표를 기반으로 열 번호 결정
            col_num = 1
            for existing_x in sorted(col_positions.keys()):
                if block["x"] > existing_x + 50:  # 50포인트 이상 차이나면 다음 열
                    col_num += 1
            
            while col_num in col_positions.values():
                col_num += 1
            
            col_positions[block["x"]] = col_num
            
            # 셀에 텍스트 입력
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = block["text"]
            
            # 폰트 크기에 따른 스타일 적용
            if block["font_size"] > 14:
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(size=10)
            
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
        ws.column_dimensions[column_letter].width = adjusted_width

def perform_xls_to_pdf_libreoffice(input_path, output_dir, timeout=120):
    """Convert XLS/XLSX to PDF using LibreOffice with enhanced error handling"""
    try:
        soffice_path = _find_soffice()
        app.logger.info(f"Using LibreOffice at: {soffice_path}")
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Build command with proper escaping
        cmd = [
            soffice_path,
            '--headless',
            '--nologo',
            '--nofirststartwizard',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            input_path
        ]
        
        app.logger.info(f"Running LibreOffice command: {' '.join(repr(arg) for arg in cmd)}")
        
        # Set environment variables for better compatibility
        env = os.environ.copy()
        env['HOME'] = '/tmp'  # Set HOME for headless operation
        env['TMPDIR'] = '/tmp'
        
        # Run LibreOffice conversion
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=env,
            cwd=output_dir  # Set working directory to output directory
        )
        
        app.logger.info(f"LibreOffice stdout: {result.stdout}")
        if result.stderr:
            app.logger.warning(f"LibreOffice stderr: {result.stderr}")
        
        if result.returncode != 0:
            app.logger.error(f"LibreOffice failed with return code {result.returncode}")
            app.logger.error(f"Command: {' '.join(cmd)}")
            app.logger.error(f"Stdout: {result.stdout}")
            app.logger.error(f"Stderr: {result.stderr}")
            return False
        
        # Check if output file was created
        input_filename = os.path.basename(input_path)
        name_without_ext = os.path.splitext(input_filename)[0]
        expected_output = os.path.join(output_dir, f"{name_without_ext}.pdf")
        
        if os.path.exists(expected_output):
            file_size = os.path.getsize(expected_output)
            app.logger.info(f"LibreOffice conversion successful - output file created: {expected_output} (size: {file_size} bytes)")
            return expected_output
        else:
            app.logger.error(f"Expected output file not found: {expected_output}")
            # List files in output directory for debugging
            try:
                files_in_output = os.listdir(output_dir)
                app.logger.info(f"Files in output directory: {files_in_output}")
            except Exception as e:
                app.logger.error(f"Could not list output directory: {e}")
            return False
            
    except subprocess.TimeoutExpired:
        app.logger.error(f"LibreOffice conversion timed out after {timeout} seconds")
        return False
    except Exception as e:
        app.logger.error(f"LibreOffice conversion failed with exception: {str(e)}")
        app.logger.exception("Full exception details:")
        return False

def _find_soffice():
    """Find LibreOffice soffice binary"""
    import glob
    
    # Common paths where LibreOffice might be installed
    possible_paths = [
        "/usr/bin/soffice",
        "/usr/lib/libreoffice/program/soffice",
        "/opt/libreoffice*/program/soffice",
        "/snap/libreoffice/current/lib/libreoffice/program/soffice"
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            app.logger.info(f"Found soffice at predefined path: {path}")
            return path
        # Handle wildcard paths
        if "*" in path:
            matches = glob.glob(path)
            if matches:
                app.logger.info(f"Found soffice at wildcard path: {matches[0]}")
                return matches[0]
    
    # Try to find using which command
    try:
        result = subprocess.run(['which', 'soffice'], capture_output=True, text=True)
        if result.returncode == 0 and result.stdout.strip():
            path = result.stdout.strip()
            app.logger.info(f"Found soffice using 'which': {path}")
            return path
    except Exception as e:
        app.logger.warning(f"Could not use 'which' to find soffice: {e}")
    
    app.logger.error("LibreOffice soffice binary not found")
    return None

@app.route("/")
def index():
    # 이 서비스 전용 텍스트/수치(템플릿에서 사용)
    page = {
        "title": "XLS to PDF Converter",
        "subtitle": "Excel 문서를 PDF로 안정적으로 변환",
        "accept": ".xls,.xlsx",
        "max_mb": os.getenv("MAX_CONTENT_LENGTH_MB", "100"),
        "service": "xls-pdf"
    }
    resp = make_response(render_template("index.html", page=page))
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Pragma"] = "no-cache"
    return resp

@app.get("/health")
def health():
    return {
        "service": "xls-pdf",
        "status": "ok"
    }, 200

@app.post("/convert-async")
def convert_async():
    f = request.files.get("file") or request.files.get("pdfFile")
    if not f: 
        return jsonify({"error":"file field is required"}), 400
    base_name = safe_base_name(f.filename)
    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}.pdf")
    f.save(in_path)

    def clamp_num(v, lo, hi, default, T=float):
        try: 
            x = T(v)
        except: 
            return default
        return max(lo, min(hi, x))

    scale = clamp_num(request.form.get("scale","1.0"), 0.2, 2.0, 1.0, float)

    JOBS[job_id] = {"status":"pending","progress":1,"message":"대기 중"}

    def run_job():
        global current_job_id
        current_job_id = job_id
        try:
            set_progress(job_id, 10, "변환 준비 중")
            
            # Adobe SDK 사용 가능하고 환경 변수가 설정된 경우에만 시도
            if ADOBE_AVAILABLE:
                try:
                    set_progress(job_id, 30, "Adobe SDK로 변환 중 (각 페이지별 시트 생성)")
                    out_path, name, ctype = perform_xlsx_conversion_adobe_with_sheets(in_path, base_name)
                    JOBS[job_id] = {"status":"done","path":out_path,"name":name,"ctype":ctype,
                                   "progress":100,"message":"Adobe SDK로 변환 완료 (각 페이지별 시트 생성)"}
                    return
                except KeyError as e:
                    app.logger.warning(f"Adobe 자격 증명 누락: {e}")
                    set_progress(job_id, 40, "Adobe 자격 증명 누락 - 폴백 변환 중")
                except Exception as e:
                    app.logger.warning(f"Adobe SDK 변환 실패: {e}")
                    set_progress(job_id, 40, "Adobe SDK 실패 - 폴백 변환 중")
            
            # LibreOffice 변환 방식 (XLS/XLSX → PDF)
            set_progress(job_id, 50, "LibreOffice로 PDF 변환 중")
            out_path = perform_xls_to_pdf_libreoffice(in_path, OUTPUTS_DIR)
            if out_path:
                name = os.path.basename(out_path)
                ctype = "application/pdf"
                JOBS[job_id] = {"status":"done","path":out_path,"name":name,"ctype":ctype,
                               "progress":100,"message":"LibreOffice로 PDF 변환 완료"}
            else:
                raise Exception("LibreOffice PDF 변환 실패")
        except Exception as e:
            app.logger.exception("convert error")
            JOBS[job_id] = {"status":"error","error":str(e),"progress":0,"message":"오류"}
        finally:
            current_job_id = None
            try: 
                os.remove(in_path)
            except: 
                pass

    executor.submit(run_job)
    return jsonify({"job_id": job_id}), 202

@app.get("/job/<job_id>")
def job_status(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    info.setdefault("progress", 0)
    info.setdefault("message","")
    return jsonify(info), 200

@app.get("/download/<job_id>")
def job_download(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    if info.get("status") != "done": 
        return jsonify({"error":"not ready"}), 409
    return send_download_memory(info["path"], info["name"], info["ctype"])

# API aliases: /api/pdf-xls/*
@app.get("/api/pdf-xls/health")
def _b1(): 
    return health()

@app.post("/api/pdf-xls/convert-async")
def _b2(): 
    return convert_async()

@app.get("/api/pdf-xls/job/<job_id>")
def _api_job_status(job_id): 
    return job_status(job_id)

@app.get("/api/pdf-xls/download/<job_id>")
def _api_job_download(job_id): 
    return job_download(job_id)

@app.post("/upload")
@app.post("/convert")
def convert_sync():
    """동기 변환 API - XLS/XLSX to PDF"""
    f = request.files.get("file") or request.files.get("xlsFile")
    if not f:
        return jsonify({"error":"파일이 없습니다"}), 400

    name = f.filename or "input.xlsx"
    if name == '':
        return jsonify({"error":"파일이 선택되지 않았습니다"}), 400
    
    ext = os.path.splitext(name)[1].lower()
    
    # XLS/XLSX 파일을 PDF로 변환하는 서비스로 변경
    ALLOWED_EXTS = {".xls", ".xlsx"}
    if ext not in ALLOWED_EXTS:
        return jsonify({"error": "XLS 또는 XLSX 파일만 업로드 가능합니다"}), 415

    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}{ext}")
    out_name = os.path.splitext(os.path.basename(name))[0] + ".pdf"
    out_path = os.path.join(OUTPUTS_DIR, f"{job_id}.pdf")
    final_path = os.path.join(OUTPUTS_DIR, out_name)

    app.logger.info(f"Starting XLS to PDF conversion: {name} -> {out_name}")
    app.logger.info(f"Input path: {in_path}")
    app.logger.info(f"Output path: {out_path}")
    app.logger.info(f"Final path: {final_path}")

    try:
        f.save(in_path)
        app.logger.info(f"File saved successfully: {in_path}")
    except Exception as e:
        app.logger.error(f"Failed to save input file: {e}")
        return jsonify({"error": f"Failed to save input file: {str(e)}"}), 500
    
    # LibreOffice 변환 시도
    conversion_success = False
    error_messages = []
    
    try:
        app.logger.info("Attempting conversion with LibreOffice...")
        result = perform_xls_to_pdf_libreoffice(in_path, OUTPUTS_DIR)
        
        if result:
            # result is the path to the created PDF file
            app.logger.info(f"LibreOffice conversion successful - output file: {result}")
            conversion_success = True
            
            # Move the file to the expected output path
            if result != out_path:
                if os.path.exists(out_path):
                    os.remove(out_path)
                os.replace(result, out_path)
                app.logger.info(f"Moved output from {result} to {out_path}")
        else:
            app.logger.error("LibreOffice conversion failed - no output file created")
            error_messages.append("LibreOffice: No output file created")
            
    except Exception as e:
        app.logger.error(f"LibreOffice conversion failed with exception: {e}")
        error_messages.append(f"LibreOffice: {str(e)}")
    
    # 변환 실패 시 오류 반환
    if not conversion_success:
        app.logger.error("All conversion methods failed")
        # 입력 파일 정리
        try:
            os.remove(in_path)
        except Exception:
            pass
        return jsonify({
            "error": "PDF conversion failed with all available methods",
            "details": error_messages
        }), 500
    
    # 파일 이동 및 최종 처리
    try:
        app.logger.info(f"Moving output file from {out_path} to {final_path}")
        
        if out_path != final_path:
            if os.path.exists(final_path):
                os.remove(final_path)
                app.logger.info(f"Removed existing final file: {final_path}")
            os.replace(out_path, final_path)
            app.logger.info(f"Moved to final path: {final_path}")
    except Exception as e:
        app.logger.error(f"Failed to move output file: {e}")
        # 원본 파일이 있다면 그것을 사용
        if os.path.exists(out_path):
            final_path = out_path
        else:
            # 입력 파일 정리
            try:
                os.remove(in_path)
            except Exception:
                pass
            return jsonify({"error": f"Failed to finalize output file: {str(e)}"}), 500
    finally:
        # 입력 파일 정리
        try:
            os.remove(in_path)
        except Exception:
            pass
    
    ctype = "application/pdf"
    return send_download_memory(final_path, out_name, ctype)

@app.post("/api/pdf-xls/convert")
def _alias_convert_sync():
    return convert_sync()

@app.get("/api/pdf-xls/job/<job_id>")
def _b3(job_id): 
    return job_status(job_id)

@app.get("/api/pdf-xls/download/<job_id>")
def _b4(job_id): 
    return job_download(job_id)

@app.errorhandler(HTTPException)
def handle_http_exc(e):
    return jsonify({"error": e.description}), e.code

@app.errorhandler(Exception)
def handle_any(e):
    app.logger.exception("Unhandled")
    return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)