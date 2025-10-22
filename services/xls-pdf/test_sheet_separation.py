import requests
import os
import tempfile
from openpyxl import load_workbook

def test_single_page_sheet_separation():
    """1페이지 PDF의 시트 분할 기능 테스트"""
    
    # 테스트용 PDF 파일 경로 (실제 테스트 파일이 있어야 함)
    test_pdf_path = "test_single_page.pdf"
    
    # 테스트 PDF가 없으면 스킵
    if not os.path.exists(test_pdf_path):
        print("테스트 PDF 파일이 없습니다. 테스트를 스킵합니다.")
        return
    
    # API 엔드포인트
    url = "http://localhost:5000/api/pdf-xls/convert"
    
    try:
        # PDF 파일 업로드 및 변환 요청
        with open(test_pdf_path, 'rb') as f:
            files = {'file': f}
            response = requests.post(url, files=files)
        
        if response.status_code == 200:
            # 응답을 임시 파일로 저장
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(response.content)
                tmp_path = tmp_file.name
            
            # Excel 파일 분석
            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames
            
            print(f"생성된 시트 목록: {sheet_names}")
            print(f"총 시트 개수: {len(sheet_names)}")
            
            # 각 시트의 내용 확인
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                row_count = ws.max_row
                print(f"시트 '{sheet_name}': {row_count}행")
                
                # 첫 몇 행의 내용 출력
                for row in range(1, min(6, row_count + 1)):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value:
                        print(f"  행 {row}: {str(cell_value)[:50]}...")
            
            # 임시 파일 정리
            os.unlink(tmp_path)
            
            # 테스트 결과 검증
            if len(sheet_names) > 1:
                print("✅ 성공: 1페이지 PDF가 여러 시트로 분할되었습니다.")
            else:
                print("⚠️ 경고: 시트가 하나만 생성되었습니다.")
                
        else:
            print(f"❌ API 요청 실패: {response.status_code}")
            print(f"응답: {response.text}")
            
    except Exception as e:
        print(f"❌ 테스트 실행 중 오류: {e}")

def test_multi_page_sheet_behavior():
    """다중 페이지 PDF의 기존 동작 확인"""
    
    # 테스트용 다중 페이지 PDF 파일 경로
    test_pdf_path = "test_multi_page.pdf"
    
    # 테스트 PDF가 없으면 스킵
    if not os.path.exists(test_pdf_path):
        print("다중 페이지 테스트 PDF 파일이 없습니다. 테스트를 스킵합니다.")
        return
    
    # API 엔드포인트
    url = "http://localhost:5000/api/pdf-xls/convert"
    
    try:
        # PDF 파일 업로드 및 변환 요청
        with open(test_pdf_path, 'rb') as f:
            files = {'file': f}
            response = requests.post(url, files=files)
        
        if response.status_code == 200:
            # 응답을 임시 파일로 저장
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(response.content)
                tmp_path = tmp_file.name
            
            # Excel 파일 분석
            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames
            
            print(f"다중 페이지 PDF - 생성된 시트 목록: {sheet_names}")
            print(f"총 시트 개수: {len(sheet_names)}")
            
            # 페이지별 시트 확인
            page_sheets = [name for name in sheet_names if name.startswith('페이지_')]
            print(f"페이지 시트 개수: {len(page_sheets)}")
            
            # 임시 파일 정리
            os.unlink(tmp_path)
            
            if len(page_sheets) > 1:
                print("✅ 성공: 다중 페이지 PDF가 페이지별로 시트 분할되었습니다.")
            else:
                print("⚠️ 경고: 예상과 다른 시트 구조입니다.")
                
        else:
            print(f"❌ API 요청 실패: {response.status_code}")
            print(f"응답: {response.text}")
            
    except Exception as e:
        print(f"❌ 테스트 실행 중 오류: {e}")

if __name__ == "__main__":
    print("=== PDF-XLS 시트 분할 기능 테스트 ===")
    print()
    
    print("1. 1페이지 PDF 시트 분할 테스트")
    test_single_page_sheet_separation()
    print()
    
    print("2. 다중 페이지 PDF 기존 동작 테스트")
    test_multi_page_sheet_behavior()
    print()
    
    print("테스트 완료")