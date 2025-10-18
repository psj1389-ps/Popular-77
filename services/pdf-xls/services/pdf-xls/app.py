import os
import sys
import io
import json
import time
import requests
import traceback
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from dotenv import load_dotenv
import fitz  # PyMuPDF
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# 환경 변수 먼저 로드
load_dotenv()

# 전역 변수 초기화
ADOBE_AVAILABLE = False
ADOBE_SDK_VERSION = None
ADOBE_IMPORT_ERROR = None

import logging

try:
    # v4 imports - 정확한 모듈 경로 사용
    from adobe.pdfservices.operation.pdf_services import PDFServices
    from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
    from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_params import ExportPDFParams
    from adobe.pdfservices.operation.pdfjobs.params.export_pdf.export_pdf_target_format import ExportPDFTargetFormat
    from adobe.pdfservices.operation.pdfjobs.jobs.export_pdf_job import ExportPDFJob
    from adobe.pdfservices.operation.pdfjobs.result.export_pdf_result import ExportPDFResult
    from adobe.pdfservices.operation.io.cloud_asset import CloudAsset
    from adobe.pdfservices.operation.io.stream_asset import StreamAsset

    ADOBE_AVAILABLE = True
    ADOBE_SDK_VERSION = "4.x"
except Exception as e:
    ADOBE_AVAILABLE = False
    ADOBE_IMPORT_ERROR = f"v4 import failed: {e}"
    logging.warning(ADOBE_IMPORT_ERROR)

# 자격증명(JSON 파일 확보)
import os, json, tempfile, base64

def ensure_adobe_creds_file() -> str:
    # 1) JSON/FILE 직접 제공
    path = os.getenv("ADOBE_CREDENTIALS_FILE_PATH") or os.getenv("PDF_SERVICES_CREDENTIALS_FILE_PATH")
    js = os.getenv("ADOBE_CREDENTIALS_JSON") or os.getenv("PDF_SERVICES_CREDENTIALS_JSON")
    if path and os.path.exists(path):
        return path
    if js:
        fd, tmp = tempfile.mkstemp(prefix="adobe_creds_", suffix=".json")
        with os.fdopen(fd, "w") as f: f.write(js)
        return tmp

    # 2) ADOBE_*로 JSON 생성 (파일/평문/B64 키 모두 지원)
    cid = os.getenv("ADOBE_CLIENT_ID")
    csec = os.getenv("ADOBE_CLIENT_SECRET")
    org = os.getenv("ADOBE_ORGANIZATION_ID")
    acct = os.getenv("ADOBE_ACCOUNT_ID") or os.getenv("ADOBE_TECHNICAL_ACCOUNT_EMAIL")
    key_path = os.getenv("ADOBE_PRIVATE_KEY_PATH")
    key_text = os.getenv("ADOBE_PRIVATE_KEY")
    key_b64 = os.getenv("ADOBE_PRIVATE_KEY_B64")

    if not (cid and csec and org and acct and (key_path or key_text or key_b64)):
        missing = [k for k, v in {
            "ADOBE_CLIENT_ID": cid,
            "ADOBE_CLIENT_SECRET": csec,
            "ADOBE_ORGANIZATION_ID": org,
            "ADOBE_ACCOUNT_ID/ADOBE_TECHNICAL_ACCOUNT_EMAIL": acct,
            "ADOBE_PRIVATE_KEY_(PATH|/|_B64)": key_path or key_text or key_b64,
        }.items() if not v]
        raise RuntimeError(f"Missing Adobe env(s): {', '.join(missing)}")

    if key_text:
        private_key = key_text
    elif key_b64:
        private_key = base64.b64decode(key_b64).decode("utf-8")
    else:
        if not os.path.exists(key_path):
            raise RuntimeError(f"Private key not found: {key_path}")
        with open(key_path, "r") as f:
            private_key = f.read()

    payload = {
        "client_credentials": {"client_id": cid, "client_secret": csec},
        "service_account_credentials": {
            "organization_id": org,
            "account_id": acct,
            "private_key": private_key
        }
    }
    fd, tmp = tempfile.mkstemp(prefix="adobe_creds_", suffix=".json")
    with os.fdopen(fd, "w") as f: json.dump(payload, f)
    logging.info("Built Adobe credentials JSON from ADOBE_* envs at %s", tmp)
    return tmp

# 추가 imports
from werkzeug.exceptions import HTTPException
import urllib.parse, tempfile, shutil, errno, logging
from uuid import uuid4
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from typing import Optional

app = Flask(__name__)
CORS(app)

# Adobe 자격 증명 (개선된 디버그 출력)
ENV_CLIENT_ID = os.getenv("ADOBE_CLIENT_ID") or os.getenv("PDF_SERVICES_CLIENT_ID") or ""
ENV_CLIENT_SECRET = os.getenv("ADOBE_CLIENT_SECRET") or os.getenv("PDF_SERVICES_CLIENT_SECRET") or ""
ENV_CREDS_JSON = os.getenv("ADOBE_CREDENTIALS_JSON") or os.getenv("PDF_SERVICES_CREDENTIALS_JSON") or ""
ENV_CREDS_FILE_PATH = os.getenv("ADOBE_CREDENTIALS_FILE_PATH") or os.getenv("PDF_SERVICES_CREDENTIALS_FILE_PATH") or ""

print("Environment Variables Check:")
print(f"ADOBE/PDF_SERVICES CLIENT_ID exists: {bool(ENV_CLIENT_ID)}")
print(f"ADOBE/PDF_SERVICES CLIENT_SECRET exists: {bool(ENV_CLIENT_SECRET)}")
print(f"ADOBE_CREDENTIALS_JSON exists: {bool(ENV_CREDS_JSON)}")
print(f"ADOBE_CREDENTIALS_FILE_PATH exists: {bool(ENV_CREDS_FILE_PATH)}")

print("=== Final Status ===")
print(f"Adobe SDK Available: {ADOBE_AVAILABLE}")
print(f"Adobe SDK Version: {ADOBE_SDK_VERSION if ADOBE_AVAILABLE else 'None'}")
print(f"Client ID configured: {bool(ENV_CLIENT_ID)}")
print(f"Client Secret configured: {bool(ENV_CLIENT_SECRET)}")

logging.basicConfig(level=logging.INFO)

# 디버그 함수: v4 모듈 네임스페이스 확인
def debug_adobe_modules():
    import importlib, pkgutil, logging
    for base in [
        "adobe.pdfservices.operation.pdfjobs.params",
        "adobe.pdfservices.operation.pdfjobs.jobs",
        "adobe.pdfservices.operation.io",
    ]:
        try:
            m = importlib.import_module(base)
            subs = [name for _, name, _ in pkgutil.iter_modules(m.__path__)]
            logging.info("%s -> %s", base, subs)
        except Exception as e:
            logging.warning("Cannot import %s: %s", base, e)

if ADOBE_AVAILABLE and ADOBE_SDK_VERSION == "4.x":
    debug_adobe_modules()

# 요청 로깅 추가
@app.before_request
def _trace():
    app.logger.info(f">>> {request.method} {request.path}")
CORS(app, resources={r"/*": {"origins": [
    "http://localhost:5173",
    "https://77-tools.xyz",
    "https://www.77-tools.xyz",
    "https://popular-77.vercel.app"
], "expose_headers": ["Content-Disposition"], "methods": ["GET","POST","OPTIONS"], "allow_headers": ["Content-Type"]}})

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)

# 환경변수에서 워커 수 설정 (기본값: 2)
max_workers = int(os.environ.get("MAX_WORKERS", 2))
executor = ThreadPoolExecutor(max_workers=max_workers)
JOBS = {}
current_job_id: Optional[str] = None

def safe_base_name(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename or "output"))[0]
    return base.replace("/", "").replace("\\", "").strip() or "output"

def set_progress(job_id, p, msg=None):
    info = JOBS.setdefault(job_id, {})
    info["progress"] = int(p)
    if msg is not None:
        info["message"] = msg

def send_download_memory(path: str, download_name: str, ctype: str):
    if not path or not os.path.exists(path):
        return jsonify({"error":"output file missing"}), 500
    with open(path, "rb") as f:
        data = f.read()
    resp = send_file(io.BytesIO(data), mimetype=ctype, as_attachment=True, download_name=download_name, conditional=False)
    resp.direct_passthrough = False
    resp.headers["Content-Length"] = str(len(data))
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{urllib.parse.quote(download_name)}"
    return resp

def _env(key: str, alt: list[str] = []):
    import os
    for k in [key, *alt]:
        v = os.environ.get(k)
        if v: return v
    raise KeyError(f"Missing env: {key}")

def adobe_context():
    if not ADOBE_AVAILABLE:
        raise RuntimeError("Adobe SDK is not installed or configured.")
    
    # 환경변수에서 자격증명 읽기 - 여러 변수명 지원
    client_id = os.getenv("ADOBE_CLIENT_ID") or os.getenv("PDF_SERVICES_CLIENT_ID")
    client_secret = os.getenv("ADOBE_CLIENT_SECRET") or os.getenv("PDF_SERVICES_CLIENT_SECRET")
    
    # 디버그 로그 추가
    logging.info(f"Adobe credentials check - Client ID exists: {bool(client_id)}, Client Secret exists: {bool(client_secret)}")
    
    if not client_id or not client_secret:
        error_msg = f"Adobe credentials not found in environment variables. CLIENT_ID: {bool(client_id)}, CLIENT_SECRET: {bool(client_secret)}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)
    
    try:
        # v4 SDK 방식: ServicePrincipalCredentials 직접 생성자 사용
        creds = ServicePrincipalCredentials(
            client_id=client_id,
            client_secret=client_secret
        )
        pdf_services = PDFServices(credentials=creds)
        logging.info("Adobe PDF Services context created successfully")
        return pdf_services
    except Exception as e:
        error_msg = f"Failed to create Adobe PDF Services context: {str(e)}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)

def _export_via_adobe(in_pdf_path: str, target: str, out_path: str):
    try:
        logging.info(f"Starting Adobe PDF export: {in_pdf_path} -> {out_path} (format: {target})")
        pdf_services = adobe_context()
        
        # v4 방식: JobFileRef 생성 및 ExportPDFJob 실행
        logging.info("Uploading PDF asset to Adobe services...")
        input_asset = pdf_services.upload(input_stream=open(in_pdf_path, 'rb'), mime_type="application/pdf")
        
        # ExportPDFParams 설정
        logging.info(f"Setting export parameters for target format: {target}")
        export_pdf_params = ExportPDFParams(target_format=ExportPDFTargetFormat[target])
        
        # ExportPDFJob 생성 및 실행
        logging.info("Creating and submitting export job...")
        export_pdf_job = ExportPDFJob(input_asset=input_asset, export_pdf_params=export_pdf_params)
        location = pdf_services.submit(export_pdf_job)
        
        logging.info(f"Job submitted, getting result from location: {location}")
        pdf_services_response = pdf_services.get_job_result(location, ExportPDFResult)
        
        # 결과 저장
        logging.info("Processing job result...")
        result_asset = pdf_services_response.get_result().get_asset()
        stream_asset = pdf_services.get_content(result_asset)
        
    except Exception as e:
        error_msg = f"Adobe PDF export failed: {str(e)}"
        logging.error(error_msg)
        logging.error(f"Traceback: {traceback.format_exc()}")
        raise RuntimeError(error_msg)
    
    if os.path.exists(out_path):
        os.remove(out_path)
    
    with open(out_path, "wb") as file:
        # stream_asset.get_input_stream()이 bytes를 반환하는 경우 처리
        input_stream = stream_asset.get_input_stream()
        if isinstance(input_stream, bytes):
            file.write(input_stream)
        else:
            file.write(input_stream.read())

# v3 변환(XLSX)
def _truthy(v: str | None) -> bool:
    return str(v).lower() in {"1", "true", "yes", "on"} if v is not None else False

def perform_xlsx_conversion_adobe(in_pdf_path: str, out_xlsx_path: str):
    # Adobe SDK 가용성 체크
    if not ADOBE_AVAILABLE or ADOBE_SDK_VERSION != "4.x":
        raise RuntimeError("Adobe v4 SDK not available")
    
    # Adobe 자격증명 체크 - 필수 환경변수가 있으면 ADOBE_DISABLED 무시
    client_id = os.getenv("ADOBE_CLIENT_ID") or os.getenv("PDF_SERVICES_CLIENT_ID")
    client_secret = os.getenv("ADOBE_CLIENT_SECRET") or os.getenv("PDF_SERVICES_CLIENT_SECRET")
    
    if not client_id or not client_secret:
        raise RuntimeError("Adobe credentials not configured")
    
    # ADOBE_DISABLED가 설정되어 있어도 자격증명이 있으면 경고만 출력하고 계속 진행
    if _truthy(os.getenv("ADOBE_DISABLED")):
        logging.warning("ADOBE_DISABLED is set but Adobe credentials are available - proceeding with conversion")

    # v4 방식으로 변환 실행
    _export_via_adobe(in_pdf_path, "XLSX", out_xlsx_path)

# 폴백 함수(scale 인자 수용)
def perform_xlsx_conversion_fallback(in_pdf_path: str, out_xlsx_path: str, scale: float = 1.0):
    import fitz, openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import io
    from PIL import Image as PILImage
    
    # 새 워크북 생성 및 기본 설정
    wb = openpyxl.Workbook()
    
    # 기본 스타일 설정
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    
    with fitz.open(in_pdf_path) as doc:
        # 1페이지 PDF인 경우 시트별로 분할
        if len(doc) == 1:
            page = doc[0]
            
            # 텍스트 추출 및 분석
            text_content = page.get_text("text")
            text_lines = [line.strip() for line in text_content.splitlines() if line.strip()]
            
            # 이미지 추출
            image_list = page.get_images(full=True)
            
            # 기본 시트 제거 후 새 시트들 생성
            wb.remove(wb.active)
            
            # 텍스트가 있는 경우 텍스트 시트 생성
            if text_lines:
                text_ws = wb.create_sheet("텍스트_내용")
                
                # 텍스트를 섹션별로 분할 (빈 줄이나 특정 패턴으로 구분)
                sections = []
                current_section = []
                
                for line in text_lines:
                    # 제목이나 헤더로 보이는 패턴 (대문자, 숫자로 시작, 특수문자 등)
                    if (len(line) < 50 and 
                        (line.isupper() or 
                         line.startswith(('1.', '2.', '3.', '4.', '5.', 'Chapter', '제', '항목', '●', '■', '-')) or
                         line.endswith(':') or
                         any(char in line for char in ['=', '*', '#']))):
                        if current_section:
                            sections.append(current_section)
                            current_section = []
                        current_section.append(line)
                    else:
                        current_section.append(line)
                
                if current_section:
                    sections.append(current_section)
                
                # 섹션이 너무 많으면 합치기 (최대 5개 섹션)
                if len(sections) > 5:
                    # 작은 섹션들을 합치기
                    merged_sections = []
                    temp_section = []
                    
                    for section in sections:
                        temp_section.extend(section)
                        if len(temp_section) >= len(text_lines) // 3:  # 전체의 1/3 정도씩
                            merged_sections.append(temp_section)
                            temp_section = []
                    
                    if temp_section:
                        if merged_sections:
                            merged_sections[-1].extend(temp_section)
                        else:
                            merged_sections.append(temp_section)
                    
                    sections = merged_sections[:5]  # 최대 5개로 제한
                
                # 섹션별로 시트 생성 또는 하나의 시트에 모든 텍스트
                if len(sections) > 1:
                    for i, section in enumerate(sections, 1):
                        section_ws = wb.create_sheet(f"섹션_{i}")
                        row = 1
                        
                        # 섹션 헤더
                        header_cell = section_ws.cell(row=row, column=1, value=f"=== 섹션 {i} ===")
                        header_cell.font = header_font
                        header_cell.alignment = Alignment(horizontal='center')
                        row += 2
                        
                        # 섹션 내용
                        for line in section:
                            cell = section_ws.cell(row=row, column=1, value=line)
                            cell.font = normal_font
                            row += 1
                        
                        # 열 너비 조정
                        _adjust_column_width(section_ws)
                else:
                    # 섹션이 하나뿐이면 전체 텍스트를 하나의 시트에
                    row = 1
                    header_cell = text_ws.cell(row=row, column=1, value="=== 전체 텍스트 ===")
                    header_cell.font = header_font
                    header_cell.alignment = Alignment(horizontal='center')
                    row += 2
                    
                    for line in text_lines:
                        cell = text_ws.cell(row=row, column=1, value=line)
                        cell.font = normal_font
                        row += 1
                    
                    # 열 너비 조정
                    _adjust_column_width(text_ws)
            
            # 이미지가 있는 경우 이미지 시트 생성
            if image_list:
                img_ws = wb.create_sheet("이미지_내용")
                row = 1
                
                # 이미지 시트 헤더
                header_cell = img_ws.cell(row=row, column=1, value="=== 이미지 목록 ===")
                header_cell.font = header_font
                header_cell.alignment = Alignment(horizontal='center')
                row += 2
                
                for img_index, img in enumerate(image_list):
                    try:
                        # 이미지 데이터 추출
                        xref = img[0]
                        pix = fitz.Pixmap(doc, xref)
                        
                        # CMYK를 RGB로 변환
                        if pix.n - pix.alpha < 4:  # GRAY or RGB
                            img_data = pix.tobytes("png")
                        else:  # CMYK
                            pix1 = fitz.Pixmap(fitz.csRGB, pix)
                            img_data = pix1.tobytes("png")
                            pix1 = None
                        
                        pix = None
                        
                        # PIL로 이미지 처리
                        pil_img = PILImage.open(io.BytesIO(img_data))
                        
                        # 이미지 크기 조정 (너무 큰 경우)
                        max_width, max_height = 400, 300
                        if pil_img.width > max_width or pil_img.height > max_height:
                            pil_img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                        
                        # 이미지를 바이트로 변환
                        img_buffer = io.BytesIO()
                        pil_img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # Excel에 이미지 추가
                        excel_img = OpenpyxlImage(img_buffer)
                        excel_img.anchor = f'B{row}'  # B열에 이미지 배치
                        img_ws.add_image(excel_img)
                        
                        # 이미지 설명 추가
                        img_ws.cell(row=row, column=1, value=f"이미지 {img_index + 1}").font = Font(italic=True, color="0066CC")
                        row += max(15, int(excel_img.height / 20))  # 이미지 높이에 따라 행 조정
                        
                    except Exception as img_error:
                        # 이미지 처리 실패 시 텍스트로 표시
                        img_ws.cell(row=row, column=1, value=f"이미지 {img_index + 1} - 처리 실패: {str(img_error)}").font = Font(italic=True, color="FF0000")
                        row += 1
                
                # 열 너비 조정
                _adjust_column_width(img_ws)
            
            # 시트가 하나도 생성되지 않은 경우 기본 시트 생성
            if len(wb.worksheets) == 0:
                default_ws = wb.create_sheet("PDF_내용")
                default_ws.cell(row=1, column=1, value="PDF에서 추출할 수 있는 내용이 없습니다.").font = normal_font
                _adjust_column_width(default_ws)
        
        else:
            # 다중 페이지 PDF인 경우 기존 방식 유지 (각 페이지를 별도 시트에)
            wb.remove(wb.active)  # 기본 시트 제거
            
            for page_num, page in enumerate(doc, start=1):
                # 각 페이지를 별도 시트로 생성
                ws = wb.create_sheet(f"페이지_{page_num}")
                row = 1
                
                # 페이지 헤더 추가
                page_header = f"=== 페이지 {page_num} ==="
                cell = ws.cell(row=row, column=1, value=page_header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                row += 1
                
                # 텍스트 추출 및 추가
                text_content = page.get_text("text")
                if text_content.strip():
                    for line in text_content.splitlines():
                        if line.strip():  # 빈 줄 제외
                            cell = ws.cell(row=row, column=1, value=line.strip())
                            cell.font = normal_font
                            row += 1
                
                # 이미지 추출 및 추가
                try:
                    image_list = page.get_images(full=True)
                    for img_index, img in enumerate(image_list):
                        try:
                            # 이미지 데이터 추출
                            xref = img[0]
                            pix = fitz.Pixmap(doc, xref)
                            
                            # CMYK를 RGB로 변환
                            if pix.n - pix.alpha < 4:  # GRAY or RGB
                                img_data = pix.tobytes("png")
                            else:  # CMYK
                                pix1 = fitz.Pixmap(fitz.csRGB, pix)
                                img_data = pix1.tobytes("png")
                                pix1 = None
                            
                            pix = None
                            
                            # PIL로 이미지 처리
                            pil_img = PILImage.open(io.BytesIO(img_data))
                            
                            # 이미지 크기 조정 (너무 큰 경우)
                            max_width, max_height = 400, 300
                            if pil_img.width > max_width or pil_img.height > max_height:
                                pil_img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                            
                            # 이미지를 바이트로 변환
                            img_buffer = io.BytesIO()
                            pil_img.save(img_buffer, format='PNG')
                            img_buffer.seek(0)
                            
                            # Excel에 이미지 추가
                            excel_img = OpenpyxlImage(img_buffer)
                            excel_img.anchor = f'B{row}'  # B열에 이미지 배치
                            ws.add_image(excel_img)
                            
                            # 이미지 설명 추가
                            ws.cell(row=row, column=1, value=f"[Image {img_index + 1}]").font = Font(italic=True, color="0066CC")
                            row += max(15, int(excel_img.height / 20))  # 이미지 높이에 따라 행 조정
                            
                        except Exception as img_error:
                            # 이미지 처리 실패 시 텍스트로 표시
                            ws.cell(row=row, column=1, value=f"[Image {img_index + 1} - 처리 실패: {str(img_error)}]").font = Font(italic=True, color="FF0000")
                            row += 1
                            
                except Exception as e:
                    # 이미지 추출 실패 시 로그만 남기고 계속 진행
                    logging.warning(f"Page {page_num} 이미지 추출 실패: {e}")
                
                # 열 너비 조정
                _adjust_column_width(ws)
    
def _adjust_column_width(worksheet):
    """워크시트의 열 너비를 자동 조정하는 헬퍼 함수"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 워크북 속성 설정 (복구 문의 방지)
    try:
        wb.properties.creator = "PDF-XLS Converter"
        wb.properties.lastModifiedBy = "PDF-XLS Converter"
        # created와 modified는 None으로 설정하지 않음 (오류 방지)
    except Exception as e:
        logging.warning(f"워크북 속성 설정 실패: {e}")
    
    # 파일 저장
    wb.save(out_xlsx_path)
    wb.close()

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

# health 응답(무엇이 보이는지 명확히)
@app.get("/health")
def health():
    import sys
    from importlib.metadata import version, PackageNotFoundError
    try:
        sdk_ver = version("pdfservices-sdk")
    except PackageNotFoundError:
        sdk_ver = None
    return {
        "python": sys.version,
        "pdfservices_sdk": sdk_ver,
        "adobe_available": ADOBE_AVAILABLE,
        "adobe_sdk_version": ADOBE_SDK_VERSION,
        "adobe_error": ADOBE_IMPORT_ERROR,
        "creds": {
            "json": bool(os.getenv("ADOBE_CREDENTIALS_JSON") or os.getenv("PDF_SERVICES_CREDENTIALS_JSON")),
            "file": bool(os.getenv("ADOBE_CREDENTIALS_FILE_PATH") or os.getenv("PDF_SERVICES_CREDENTIALS_FILE_PATH")),
            "cid": bool(os.getenv("ADOBE_CLIENT_ID")),
            "csecret": bool(os.getenv("ADOBE_CLIENT_SECRET")),
            "org": bool(os.getenv("ADOBE_ORGANIZATION_ID")),
            "acct": bool(os.getenv("ADOBE_ACCOUNT_ID") or os.getenv("ADOBE_TECHNICAL_ACCOUNT_EMAIL")),
            "key_file_exists": os.path.exists(os.getenv("ADOBE_PRIVATE_KEY_PATH", "")),
            "key_text": bool(os.getenv("ADOBE_PRIVATE_KEY")),
            "key_b64": bool(os.getenv("ADOBE_PRIVATE_KEY_B64")),
            "disabled": os.getenv("ADOBE_DISABLED"),
        },
    }

@app.post("/convert-async")
def convert_async():
    f = request.files.get("file") or request.files.get("pdfFile")
    if not f: 
        return jsonify({"error":"file field is required"}), 400
    base_name = safe_base_name(f.filename)
    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}.pdf")
    f.save(in_path)

    def clamp_num(v, lo, hi, default, T=float):
        try: 
            x = T(v)
        except: 
            return default
        return max(lo, min(hi, x))

    scale = clamp_num(request.form.get("scale","1.0"), 0.2, 2.0, 1.0, float)

    JOBS[job_id] = {"status":"pending","progress":1,"message":"대기 중"}

    def run_job():
        global current_job_id
        current_job_id = job_id
        try:
            set_progress(job_id, 10, "변환 준비 중")
            final_name = f"{base_name}.xlsx"
            final_path = os.path.join(OUTPUTS_DIR, final_name)
            try:
                perform_xlsx_conversion_adobe(in_path, final_path)
            except Exception as e:
                app.logger.exception("Adobe export failed; fallback to image-based.")
                perform_xlsx_conversion_fallback(in_path, final_path)
            out_path, name, ctype = final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            JOBS[job_id] = {"status":"done","path":out_path,"name":name,"ctype":ctype,
                           "progress":100,"message":"완료"}
        except Exception as e:
            app.logger.exception("convert error")
            JOBS[job_id] = {"status":"error","error":str(e),"progress":0,"message":"오류"}
        finally:
            current_job_id = None
            try: 
                os.remove(in_path)
            except: 
                pass

    executor.submit(run_job)
    return jsonify({"job_id": job_id}), 202

@app.get("/job/<job_id>")
def job_status(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    info.setdefault("progress", 0)
    info.setdefault("message","")
    return jsonify(info), 200

@app.get("/download/<job_id>")
def job_download(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    if info.get("status") != "done": 
        return jsonify({"error":"not ready"}), 409
    return send_download_memory(info["path"], info["name"], info["ctype"])

# API aliases: /api/pdf-xls/*
@app.post("/convert")
def convert_sync():
    f = request.files.get("file") or request.files.get("pdfFile")
    if not f:
        return jsonify({"error":"file field is required"}), 400

    base_name = safe_base_name(f.filename)
    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}.pdf")
    f.save(in_path)

    quality = request.form.get("quality", "low")  # 수신만, 변환에는 사용하지 않음
    def clamp_num(v, lo, hi, default, T=float):
        try: 
            x = T(v)
        except: 
            return default
        return max(lo, min(hi, x))
    scale = clamp_num(request.form.get("scale","1.0"), 0.2, 2.0, 1.0, float)

    # 라우팅/실행 흐름은 기존처럼 Adobe → 실패시 폴백
    def convert_pdf_to_xlsx(input_pdf_path, output_xlsx_path, scale: float = 1.0):
        import logging
        if ADOBE_AVAILABLE and ADOBE_SDK_VERSION == "4.x":
            try:
                perform_xlsx_conversion_adobe(input_pdf_path, output_xlsx_path)
                return
            except Exception as e:
                logging.warning("Adobe v4 conversion failed; fallback. err=%s", e, exc_info=True)
        # 최종 폴백
        perform_xlsx_conversion_fallback(input_pdf_path, output_xlsx_path, scale=scale)

    try:
        # 1. Adobe API를 먼저 시도합니다.
        app.logger.info("Attempting conversion with Adobe API...")
        final_name = f"{base_name}.xlsx"
        final_path = os.path.join(OUTPUTS_DIR, final_name)
        
        convert_pdf_to_xlsx(in_path, final_path, scale=scale)
        
        out_path, name, ctype = final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    except Exception as e:
        # 2. 어떤 이유로든 실패하면 (자격증명 오류, Adobe 서버 문제 등)
        app.logger.exception("Adobe export failed; falling back to image-based conversion.")
        # 3. 기존의 이미지 방식으로 변환을 시도합니다.
        perform_xlsx_conversion_fallback(in_path, final_path, scale=scale)
        out_path, name, ctype = final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    finally:
        try: 
            os.remove(in_path)
        except: 
            pass
    
    return send_download_memory(out_path, name, ctype)

# /api aliases (frontend uses /api paths only)
@app.route("/api/pdf-xls/health", methods=["GET", "HEAD"])
def _xls_a_health(): return health()

@app.route("/api/pdf-xls/convert", methods=["POST", "OPTIONS"])
def _xls_a_convert_sync(): return convert_sync()

@app.route("/api/pdf-xls/convert-async", methods=["POST", "OPTIONS"])
def _xls_a_convert_async(): return convert_async()

@app.route("/api/pdf-xls/job/<job_id>", methods=["GET", "HEAD"])
def _xls_a_job(job_id): return job_status(job_id)

@app.route("/api/pdf-xls/download/<job_id>", methods=["GET", "HEAD"])
def _xls_a_download(job_id): return job_download(job_id)

# Protection for wrong methods
@app.route("/api/pdf-xls/convert-async", methods=["GET", "HEAD"])
def _xls_a_convert_async_wrong_method():
    return jsonify({"error": "use POST for /convert-async"}), 405

@app.errorhandler(HTTPException)
def handle_http_exc(e):
    return jsonify({"error": e.description}), e.code

@app.errorhandler(Exception)
def handle_any(e):
    app.logger.exception("Unhandled")
    return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    # 환경변수에서 포트 설정
    port = int(os.environ.get("PORT", 5000))
    
    # 환경변수에서 호스트 설정 (기본값: 0.0.0.0)
    host = os.environ.get("HOST", "0.0.0.0")
    
    # 환경변수에서 디버그 모드 설정 (기본값: False)
    debug = os.environ.get("DEBUG", "false").lower() in ("true", "1", "yes", "on")
    
    app.run(host=host, port=port, debug=debug)