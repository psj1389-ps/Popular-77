from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
import io, os, urllib.parse, tempfile, shutil, errno, logging
from uuid import uuid4
from concurrent.futures import ThreadPoolExecutor
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional
import re

# Adobe PDF Services SDK imports
ADOBE_AVAILABLE = False
try:
    from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
    from adobe.pdfservices.operation.execution_context import ExecutionContext
    from adobe.pdfservices.operation.io.file_ref import FileRef
    from adobe.pdfservices.operation.pdfops.export_pdf_operation import ExportPDFOperation
    from adobe.pdfservices.operation.pdfops.options.export_pdf_options import ExportPDFTargetFormat
    ADOBE_AVAILABLE = True
    logging.info("Adobe PDF Services SDK loaded successfully")
except ImportError as e:
    logging.warning(f"Adobe PDF Services SDK not available: {e}")
    logging.info("Service will use fallback image-based conversion")

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
CORS(app, resources={r"/*": {"origins": [
    "http://localhost:5173",
    "https://77-tools.xyz",
    "https://www.77-tools.xyz",
    "https://popular-77.vercel.app"
], "expose_headers": ["Content-Disposition"], "methods": ["GET","POST","OPTIONS"], "allow_headers": ["Content-Type"]}})

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)

executor = ThreadPoolExecutor(max_workers=2)
JOBS = {}
current_job_id: Optional[str] = None

def safe_base_name(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename or "output"))[0]
    return base.replace("/", "").replace("\\", "").strip() or "output"

def set_progress(job_id, p, msg=None):
    info = JOBS.setdefault(job_id, {})
    info["progress"] = int(p)
    if msg is not None:
        info["message"] = msg

def send_download_memory(path: str, download_name: str, ctype: str):
    if not path or not os.path.exists(path):
        return jsonify({"error":"output file missing"}), 500
    with open(path, "rb") as f:
        data = f.read()
    resp = send_file(io.BytesIO(data), mimetype=ctype, as_attachment=True, download_name=download_name, conditional=False)
    resp.direct_passthrough = False
    resp.headers["Content-Length"] = str(len(data))
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{urllib.parse.quote(download_name)}"
    return resp

def _env(key: str, alt: list[str] = []):
    import os
    for k in [key, *alt]:
        v = os.environ.get(k)
        if v: return v
    raise KeyError(f"Missing env: {key}")

def adobe_context():
    # ADOBE_* 또는 PDF_SERVICES_* 둘 다 허용
    client_id = os.environ.get("ADOBE_CLIENT_ID") or os.environ.get("PDF_SERVICES_CLIENT_ID")
    client_secret = os.environ.get("ADOBE_CLIENT_SECRET") or os.environ.get("PDF_SERVICES_CLIENT_SECRET")
    org_id = os.environ.get("ADOBE_ORGANIZATION_ID") or os.environ.get("PDF_SERVICES_ORGANIZATION_ID")
    account_id = os.environ.get("ADOBE_ACCOUNT_ID") or os.environ.get("PDF_SERVICES_ACCOUNT_ID")
    if not client_id or not client_secret:
        raise KeyError("Adobe credentials missing: ADOBE_CLIENT_ID / ADOBE_CLIENT_SECRET")
    if org_id and account_id:
        creds = ServicePrincipalCredentials(client_id=client_id, client_secret=client_secret,
                                          organization_id=org_id, account_id=account_id)
    else:
        creds = ServicePrincipalCredentials(client_id=client_id, client_secret=client_secret)
    return ExecutionContext.create(creds)

def _export_via_adobe(in_pdf_path: str, target: str, out_path: str):
    ctx = adobe_context()
    op = ExportPDFOperation.create_new(ExportPDFTargetFormat[target])  # "XLSX"
    op.set_input(FileRef.create_from_local_file(in_pdf_path))
    result = op.execute(ctx)
    if os.path.exists(out_path):
        os.remove(out_path)
    result.save_as(out_path)

def perform_xlsx_conversion_adobe(in_path: str, base_name: str):
    final_name = f"{base_name}.xlsx"
    final_path = os.path.join(OUTPUTS_DIR, final_name)
    _export_via_adobe(in_path, "XLSX", final_path)
    return final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def has_extractable_text(page):
    """PDF 페이지에서 추출 가능한 텍스트가 있는지 확인"""
    text = page.get_text().strip()
    # 텍스트가 있고, 단순히 공백이나 특수문자만이 아닌 경우
    return len(text) > 10 and bool(re.search(r'[a-zA-Z가-힣0-9]', text))

def extract_text_to_excel(page, ws):
    """PDF 페이지에서 텍스트를 추출하여 Excel 셀에 배치"""
    # 텍스트 블록 추출 (위치 정보 포함)
    blocks = page.get_text("dict")["blocks"]
    
    # 텍스트 블록들을 y 좌표로 정렬 (위에서 아래로)
    text_blocks = []
    for block in blocks:
        if "lines" in block:  # 텍스트 블록인 경우
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        text_blocks.append({
                            "text": text,
                            "x": span["bbox"][0],
                            "y": span["bbox"][1],
                            "font_size": span["size"]
                        })
    
    # y 좌표로 정렬 후 x 좌표로 정렬
    text_blocks.sort(key=lambda x: (x["y"], x["x"]))
    
    # 행별로 그룹화 (비슷한 y 좌표끼리)
    rows = []
    current_row = []
    current_y = None
    tolerance = 5  # y 좌표 허용 오차
    
    for block in text_blocks:
        if current_y is None or abs(block["y"] - current_y) <= tolerance:
            current_row.append(block)
            current_y = block["y"] if current_y is None else current_y
        else:
            if current_row:
                rows.append(current_row)
            current_row = [block]
            current_y = block["y"]
    
    if current_row:
        rows.append(current_row)
    
    # Excel에 텍스트 배치
    for row_idx, row_blocks in enumerate(rows, 1):
        # 행 내에서 x 좌표로 정렬
        row_blocks.sort(key=lambda x: x["x"])
        
        # 열 위치 계산 (x 좌표 기반)
        col_positions = {}
        for i, block in enumerate(row_blocks):
            # x 좌표를 기반으로 열 번호 결정
            col_num = 1
            for existing_x in sorted(col_positions.keys()):
                if block["x"] > existing_x + 50:  # 50포인트 이상 차이나면 다음 열
                    col_num += 1
            
            while col_num in col_positions.values():
                col_num += 1
            
            col_positions[block["x"]] = col_num
            
            # 셀에 텍스트 입력
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = block["text"]
            
            # 폰트 크기에 따른 스타일 적용
            if block["font_size"] > 14:
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(size=10)
            
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
        ws.column_dimensions[column_letter].width = adjusted_width

def perform_xlsx_conversion(in_path: str, base_name: str, scale: float = 1.0):
    """
    PDF → XLSX (텍스트 추출 우선, 필요시 이미지 폴백)
    """
    with tempfile.TemporaryDirectory(dir=OUTPUTS_DIR) as tmp:
        doc = fitz.open(in_path)
        mat = fitz.Matrix(scale, scale)
        wb = Workbook()
        # openpyxl 기본 시트를 첫 페이지로 사용, 나머지는 추가
        first_sheet = True

        for i in range(doc.page_count):
            set_progress(current_job_id, 10 + int(80 * (i + 1) / doc.page_count), f"페이지 {i+1}/{doc.page_count} 처리 중")
            page = doc.load_page(i)
            
            if first_sheet:
                ws = wb.active
                ws.title = f"Page_{i+1:02d}"
                first_sheet = False
            else:
                ws = wb.create_sheet(title=f"Page_{i+1:02d}")
            
            # 텍스트 추출 시도
            if has_extractable_text(page):
                try:
                    extract_text_to_excel(page, ws)
                    # 텍스트 추출 성공 시 다음 페이지로
                    continue
                except Exception as e:
                    app.logger.warning(f"텍스트 추출 실패 (페이지 {i+1}): {e}")
            
            # 텍스트 추출 실패 시 이미지로 폴백
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_path = os.path.join(tmp, f"{base_name}_{i+1:02d}.png")
            pix.save(img_path)
            
            # 기존 셀 내용 삭제 (텍스트 추출 실패한 경우)
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
            
            xlimg = XLImage(img_path)
            ws.add_image(xlimg, "A1")

        final_name = f"{base_name}.xlsx"
        final_path = os.path.join(OUTPUTS_DIR, final_name)
        if os.path.exists(final_path): 
            os.remove(final_path)
        wb.save(final_path)
        return final_path, final_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

@app.route("/")
def index():
    return render_template("index.html")

@app.get("/health")
def health():
    return "ok", 200

@app.post("/convert-async")
def convert_async():
    f = request.files.get("file") or request.files.get("pdfFile")
    if not f: 
        return jsonify({"error":"file field is required"}), 400
    base_name = safe_base_name(f.filename)
    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}.pdf")
    f.save(in_path)

    def clamp_num(v, lo, hi, default, T=float):
        try: 
            x = T(v)
        except: 
            return default
        return max(lo, min(hi, x))

    scale = clamp_num(request.form.get("scale","1.0"), 0.2, 2.0, 1.0, float)

    JOBS[job_id] = {"status":"pending","progress":1,"message":"대기 중"}

    def run_job():
        global current_job_id
        current_job_id = job_id
        try:
            set_progress(job_id, 10, "변환 준비 중")
            try:
                set_progress(job_id, 30, "Adobe로 변환 중")
                out_path, name, ctype = perform_xlsx_conversion_adobe(in_path, base_name)
            except Exception as e:
                app.logger.exception("Adobe export failed; fallback to image-based.")
                set_progress(job_id, 50, "이미지 기반 폴백 변환 중")
                out_path, name, ctype = perform_xlsx_conversion(in_path, base_name, scale=scale)  # 기존 이미지 방식
            JOBS[job_id] = {"status":"done","path":out_path,"name":name,"ctype":ctype,
                           "progress":100,"message":"완료"}
        except Exception as e:
            app.logger.exception("convert error")
            JOBS[job_id] = {"status":"error","error":str(e),"progress":0,"message":"오류"}
        finally:
            current_job_id = None
            try: 
                os.remove(in_path)
            except: 
                pass

    executor.submit(run_job)
    return jsonify({"job_id": job_id}), 202

@app.get("/job/<job_id>")
def job_status(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    info.setdefault("progress", 0)
    info.setdefault("message","")
    return jsonify(info), 200

@app.get("/download/<job_id>")
def job_download(job_id):
    info = JOBS.get(job_id)
    if not info: 
        return jsonify({"error":"job not found"}), 404
    if info.get("status") != "done": 
        return jsonify({"error":"not ready"}), 409
    return send_download_memory(info["path"], info["name"], info["ctype"])

# API aliases: /api/pdf-xls/*
@app.get("/api/pdf-xls/health")
def _b1(): 
    return health()

@app.post("/api/pdf-xls/convert-async")
def _b2(): 
    return convert_async()

@app.get("/api/pdf-xls/job/<job_id>")
def _api_job_status(job_id): 
    return job_status(job_id)

@app.get("/api/pdf-xls/download/<job_id>")
def _api_job_download(job_id): 
    return job_download(job_id)

@app.post("/convert")
def convert_sync():
    f = request.files.get("file") or request.files.get("pdfFile")
    if not f:
        return jsonify({"error":"file field is required"}), 400

    base_name = safe_base_name(f.filename)
    job_id = uuid4().hex
    in_path = os.path.join(UPLOADS_DIR, f"{job_id}.pdf")
    f.save(in_path)

    quality = request.form.get("quality", "low")  # 수신만, 변환에는 사용하지 않음
    def clamp_num(v, lo, hi, default, T=float):
        try: 
            x = T(v)
        except: 
            return default
        return max(lo, min(hi, x))
    scale = clamp_num(request.form.get("scale","1.0"), 0.2, 2.0, 1.0, float)

    try:
        out_path, name, ctype = perform_xlsx_conversion_adobe(in_path, base_name)
    except Exception:
        out_path, name, ctype = perform_xlsx_conversion(in_path, base_name, scale=scale)
    finally:
        try: 
            os.remove(in_path)
        except: 
            pass
    
    return send_download_memory(out_path, name, ctype)

@app.post("/api/pdf-xls/convert")
def _alias_convert_sync():
    return convert_sync()

@app.get("/api/pdf-xls/job/<job_id>")
def _b3(job_id): 
    return job_status(job_id)

@app.get("/api/pdf-xls/download/<job_id>")
def _b4(job_id): 
    return job_download(job_id)

@app.errorhandler(HTTPException)
def handle_http_exc(e):
    return jsonify({"error": e.description}), e.code

@app.errorhandler(Exception)
def handle_any(e):
    app.logger.exception("Unhandled")
    return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)